from openpyxl import load_workbook


class Read():
    """ Read """
    def __init__(self, filename, names_range, shifts_range, skip_blank_cell=True):
        """
        Arguments:
        - `filename`:
        - `names_range`:
        - `shifts_range`:
        - `skip_blank_cell`:
        """
        self._filename = filename
        self._names_range = names_range
        self._shifts_range = shifts_range
        self._skip_blank_cell = skip_blank_cell

        self._load()

    def _load(self):
        """
        """
        self._wb = load_workbook(self._filename)
        self._ws = self._wb.get_active_sheet()

    def get_names(self):
        """
        Arguments:
        - `self`:
        """
        names = [c[0].value for c in self._ws.range(self._names_range)]
        return [name for name in names if
                not self._skip_blank_cell or name]

    def get_shifts(self):
        """
        Arguments:
        - `self`:
        """
        shifts = [[c.value for c in row]
                  for row in self._ws.range(self._shifts_range)]
        return [shift for shift in shifts
                if not self._skip_blank_cell or any(shift)]


def load_roster(file_name, names_range, shifts_range, skip_blank_cell=True):
    """
    Arguments:
    - `file_name`: String -- File name of roster.xlsx
    - `names_range`: String -- Range of names. -> "A3:A10"
    - `shifts_range`: String -- Range of shifts.
    - `skip_blank_cell`: Bool -- Whether to skip empty cells.
    return: ([String]     # names
            , [[String]]) # shifts
    """
    wb = load_workbook(file_name)
    ws = wb.get_active_sheet()
    names = [c[0].value for c in ws.range(names_range)]
    shifts = [[c.value for c in row] for row in ws.range(shifts_range)]
    if skip_blank_cell:
        names = [name for name in names if name]
        shifts = [shift for shift in shifts if any(shift)]
    return (names, shifts)
