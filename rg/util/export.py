from openpyxl import load_workbook

__all__ = ["to_xlsx"]


def to_xlsx(roster, template, export_filename, date):
    wb = load_workbook(template)
    ws = wb.get_active_sheet()
    wsls = ws.range("C7:AG15")
    ws.cell("B2").value = date[0]
    ws.cell("C2").value = date[1]
    for i, shift in enumerate(roster):
        ii = i if i < 4 else i + 2
        for j, day in enumerate(shift):
            wsls[ii][j].value = day.work
    wb.save(export_filename)
